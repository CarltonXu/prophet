# Copyright (c) 2021 OnePro Cloud Ltd.
#
#   prophet is licensed under Mulan PubL v2.

"""Hosts API"""

from flask import Blueprint, request, jsonify, Response, send_file
from flask_jwt_extended import jwt_required
from utils.jwt import get_current_user_id
from sqlalchemy import or_, and_
from sqlalchemy.orm import selectinload
import json
import csv
import io

from models import Host, HostCredential, HostDetail, HostTag, HostTagRelation, HostRelationship, CollectionTask, db
from tasks.collector import collect_hosts_task
from utils.encryption import encrypt_password, decrypt_password
from utils.decorators import validate_json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

FIELD_CATEGORIES = {
    'basic': {
        'label': '基本信息',
        'label_key': 'hosts.exportCategories.basic'
    },
    'hardware': {
        'label': '硬件信息',
        'label_key': 'hosts.exportCategories.hardware'
    },
    'storage': {
        'label': '存储信息',
        'label_key': 'hosts.exportCategories.storage'
    },
    'network': {
        'label': '网络信息',
        'label_key': 'hosts.exportCategories.network'
    },
    'virtualization': {
        'label': '虚拟化信息',
        'label_key': 'hosts.exportCategories.virtualization'
    },
    'status': {
        'label': '采集状态',
        'label_key': 'hosts.exportCategories.status'
    }
}


def _bytes_to_gb(value):
    if not value:
        return ''
    try:
        return round(value / (1024 ** 3), 2)
    except Exception:
        return ''


FIELD_DEFINITIONS = {
    'platform_type': {
        'header': '平台类型',
        'label': '平台类型',
        'label_key': 'hosts.exportColumns.platformType',
        'category': 'basic',
        'width': 12,
        'align': 'center'
    },
    'hostname': {
        'header': '主机名',
        'label': '主机名',
        'label_key': 'hosts.exportColumns.hostname',
        'category': 'basic',
        'width': 22
    },
    'vmware_host': {
        'header': 'VMware主机名',
        'label': 'VMware主机名',
        'label_key': 'hosts.exportColumns.vmwareHostName',
        'category': 'virtualization',
        'width': 24
    },
    'ip': {
        'header': 'IP',
        'label': 'IP',
        'label_key': 'hosts.exportColumns.ipAddress',
        'category': 'basic',
        'width': 18
    },
    'mac': {
        'header': 'Mac',
        'label': 'Mac',
        'label_key': 'hosts.exportColumns.macAddress',
        'category': 'basic',
        'width': 20
    },
    'os_type': {
        'header': '操作系统类型',
        'label': '操作系统类型',
        'label_key': 'hosts.exportColumns.osType',
        'category': 'basic',
        'width': 18
    },
    'os_version': {
        'header': '操作系统版本',
        'label': '操作系统版本',
        'label_key': 'hosts.exportColumns.osVersion',
        'category': 'basic',
        'width': 28
    },
    'os_bit': {
        'header': '操作系统位数',
        'label': '操作系统位数',
        'label_key': 'hosts.exportColumns.osBit',
        'category': 'basic',
        'width': 12,
        'align': 'center'
    },
    'os_kernel': {
        'header': '操作系统内核',
        'label': '操作系统内核',
        'label_key': 'hosts.exportColumns.osKernel',
        'category': 'basic',
        'width': 20
    },
    'boot_type': {
        'header': '启动方式',
        'label': '启动方式',
        'label_key': 'hosts.exportColumns.bootType',
        'category': 'basic',
        'width': 12,
        'align': 'center'
    },
    'cpu_info': {
        'header': 'CPU',
        'label': 'CPU',
        'label_key': 'hosts.exportColumns.cpuInfo',
        'category': 'hardware',
        'width': 32
    },
    'cpu_cores': {
        'header': 'CPU核数',
        'label': 'CPU核数',
        'label_key': 'hosts.exportColumns.cpuCores',
        'category': 'hardware',
        'width': 12,
        'align': 'center'
    },
    'memory_label': {
        'header': '内存',
        'label': '内存',
        'label_key': 'hosts.exportColumns.memoryLabel',
        'category': 'hardware',
        'width': 18
    },
    'memory_total_gb': {
        'header': '总内存(GB)',
        'label': '总内存(GB)',
        'label_key': 'hosts.exportColumns.memoryTotal',
        'category': 'hardware',
        'width': 16,
        'align': 'center'
    },
    'memory_free_gb': {
        'header': '剩余内存(GB)',
        'label': '剩余内存(GB)',
        'label_key': 'hosts.exportColumns.memoryFree',
        'category': 'hardware',
        'width': 18,
        'align': 'center'
    },
    'disk_count': {
        'header': '磁盘数量',
        'label': '磁盘数量',
        'label_key': 'hosts.exportColumns.diskCount',
        'category': 'storage',
        'width': 12,
        'align': 'center'
    },
    'disk_total_size_gb': {
        'header': '磁盘总容量(GB)',
        'label': '磁盘总容量(GB)',
        'label_key': 'hosts.exportColumns.diskTotalSize',
        'category': 'storage',
        'width': 18,
        'align': 'center'
    },
    'disks': {
        'header': '磁盘信息',
        'label': '磁盘信息',
        'label_key': 'hosts.exportColumns.disks',
        'category': 'storage',
        'width': 55,
        'wrap': True
    },
    'partitions': {
        'header': '分区信息',
        'label': '分区信息',
        'label_key': 'hosts.exportColumns.partitions',
        'category': 'storage',
        'width': 55,
        'wrap': True
    },
    'network_count': {
        'header': '网卡数量',
        'label': '网卡数量',
        'label_key': 'hosts.exportColumns.networkCount',
        'category': 'network',
        'width': 12,
        'align': 'center'
    },
    'networks': {
        'header': '网卡信息',
        'label': '网卡信息',
        'label_key': 'hosts.exportColumns.networkInterfaces',
        'category': 'network',
        'width': 60,
        'wrap': True
    },
    'vt_platform': {
        'header': '虚拟化类型',
        'label': '虚拟化类型',
        'label_key': 'hosts.exportColumns.vtPlatform',
        'category': 'virtualization',
        'width': 20
    },
    'vt_platform_ver': {
        'header': '虚拟化版本',
        'label': '虚拟化版本',
        'label_key': 'hosts.exportColumns.vtPlatformVersion',
        'category': 'virtualization',
        'width': 20
    },
    'esxi_host': {
        'header': 'ESXi服务器',
        'label': 'ESXi服务器',
        'label_key': 'hosts.exportColumns.esxiHost',
        'category': 'virtualization',
        'width': 25
    },
    'collection_status': {
        'header': '采集状态',
        'label': '采集状态',
        'label_key': 'hosts.exportColumns.collectionStatus',
        'category': 'status',
        'width': 16,
        'align': 'center'
    },
    'last_collected_at': {
        'header': '最后采集时间',
        'label': '最后采集时间',
        'label_key': 'hosts.exportColumns.lastCollectedAt',
        'category': 'status',
        'width': 22
    }
}

EXPORT_TEMPLATES = {
    'summary': {
        'name': '概要模板',
        'name_key': 'hosts.exportTemplates.summary',
        'description': '包含关键基础信息，适合快速预览',
        'description_key': 'hosts.exportTemplates.summaryDesc',
        'fields': [
            'platform_type', 'hostname', 'ip', 'os_type', 'cpu_info', 'cpu_cores',
            'memory_total_gb', 'disk_total_size_gb', 'vt_platform', 'collection_status'
        ]
    },
    'detailed': {
        'name': '详细模板',
        'name_key': 'hosts.exportTemplates.detailed',
        'description': '包含硬件、存储和网络详细信息',
        'description_key': 'hosts.exportTemplates.detailedDesc',
        'fields': [
            'platform_type', 'hostname', 'vmware_host', 'ip', 'mac',
            'os_type', 'os_version', 'os_bit', 'os_kernel', 'boot_type',
            'cpu_info', 'cpu_cores', 'memory_label', 'memory_total_gb', 'memory_free_gb',
            'disk_count', 'disk_total_size_gb', 'disks', 'partitions',
            'network_count', 'networks', 'vt_platform', 'vt_platform_ver', 'esxi_host',
            'collection_status', 'last_collected_at'
        ]
    },
    'all': {
        'name': '全部字段',
        'name_key': 'hosts.exportTemplates.all',
        'description': '导出所有支持的字段',
        'description_key': 'hosts.exportTemplates.allDesc',
        'fields': list(FIELD_DEFINITIONS.keys())
    }
}

DEFAULT_EXPORT_TEMPLATE = 'summary'

HOST_EXPORT_OPTIONS = [
    selectinload(Host.disks),
    selectinload(Host.partitions),
    selectinload(Host.network_interfaces),
    selectinload(Host.platform),
]

HOST_TEMPLATE_EXPORT_OPTIONS = [
    selectinload(Host.credentials),
    selectinload(Host.tags),
]

HOST_IMPORT_TEMPLATE_COLUMNS = [
    {
        'key': 'ip',
        'label': 'IP',
        'label_key': 'hosts.importFields.ip',
        'description_key': 'hosts.importFields.ipDesc',
        'required': True,
        'example': '192.168.1.10'
    },
    {
        'key': 'hostname',
        'label': 'Hostname',
        'label_key': 'hosts.importFields.hostname',
        'description_key': 'hosts.importFields.hostnameDesc',
        'required': False,
        'example': 'web-server-01'
    },
    {
        'key': 'mac',
        'label': 'MAC',
        'label_key': 'hosts.importFields.mac',
        'description_key': 'hosts.importFields.macDesc',
        'required': False,
        'example': '00:16:3e:3d:4f:5a'
    },
    {
        'key': 'vendor',
        'label': 'Vendor',
        'label_key': 'hosts.importFields.vendor',
        'description_key': 'hosts.importFields.vendorDesc',
        'required': False,
        'example': 'Dell'
    },
    {
        'key': 'os_type',
        'label': 'OS Type',
        'label_key': 'hosts.importFields.osType',
        'description_key': 'hosts.importFields.osTypeDesc',
        'required': False,
        'example': 'CentOS'
    },
    {
        'key': 'os_version',
        'label': 'OS Version',
        'label_key': 'hosts.importFields.osVersion',
        'description_key': 'hosts.importFields.osVersionDesc',
        'required': False,
        'example': '7.9'
    },
    {
        'key': 'device_type',
        'label': 'Device Type',
        'label_key': 'hosts.importFields.deviceType',
        'description_key': 'hosts.importFields.deviceTypeDesc',
        'required': False,
        'example': 'host'
    },
    {
        'key': 'is_physical',
        'label': 'Is Physical',
        'label_key': 'hosts.importFields.isPhysical',
        'description_key': 'hosts.importFields.isPhysicalDesc',
        'required': False,
        'example': '是/否'
    },
    {
        'key': 'vt_platform',
        'label': 'Virtualization Platform',
        'label_key': 'hosts.importFields.vtPlatform',
        'description_key': 'hosts.importFields.vtPlatformDesc',
        'required': False,
        'example': 'VMware'
    },
    {
        'key': 'vt_platform_ver',
        'label': 'Virtualization Version',
        'label_key': 'hosts.importFields.vtPlatformVer',
        'description_key': 'hosts.importFields.vtPlatformVerDesc',
        'required': False,
        'example': '7.0'
    },
    {
        'key': 'username',
        'label': 'Username',
        'label_key': 'hosts.importFields.username',
        'description_key': 'hosts.importFields.usernameDesc',
        'required': False,
        'example': 'root'
    },
    {
        'key': 'password',
        'label': 'Password',
        'label_key': 'hosts.importFields.password',
        'description_key': 'hosts.importFields.passwordDesc',
        'required': False,
        'example': ''
    },
    {
        'key': 'ssh_port',
        'label': 'SSH Port',
        'label_key': 'hosts.importFields.sshPort',
        'description_key': 'hosts.importFields.sshPortDesc',
        'required': False,
        'example': '22'
    },
    {
        'key': 'key_path',
        'label': 'Key Path',
        'label_key': 'hosts.importFields.keyPath',
        'description_key': 'hosts.importFields.keyPathDesc',
        'required': False,
        'example': '/home/user/.ssh/id_rsa'
    },
    {
        'key': 'tags',
        'label': 'Tags',
        'label_key': 'hosts.importFields.tags',
        'description_key': 'hosts.importFields.tagsDesc',
        'required': False,
        'example': '生产,web'
    },
]


bp = Blueprint('hosts', __name__)


@bp.route('', methods=['GET'])
@jwt_required()
def get_hosts():
    """Get host list with filtering and pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # Filters
    search = request.args.get('search')
    search_field = request.args.get('search_field', 'all')  # all, ip, hostname, mac, vendor
    os_type = request.args.get('os_type')
    device_type = request.args.get('device_type')
    is_physical = request.args.get('is_physical')
    platform_id = request.args.get('platform_id')
    tag_id = request.args.get('tag_id')
    collection_status = request.args.get('collection_status')
    source = request.args.get('source')
    
    query = Host.query.filter(Host.deleted_at == None)
    
    if search:
        if search_field == 'ip':
            query = query.filter(Host.ip.contains(search))
        elif search_field == 'hostname':
            query = query.filter(Host.hostname.contains(search))
        elif search_field == 'mac':
            query = query.filter(Host.mac.contains(search))
        elif search_field == 'vendor':
            query = query.filter(Host.vendor.contains(search))
        else:
            # Default: search all fields
            query = query.filter(
                or_(
                    Host.hostname.contains(search),
                    Host.ip.contains(search),
                    Host.mac.contains(search),
                    Host.vendor.contains(search)
                )
            )
    
    if os_type:
        query = query.filter_by(os_type=os_type)
    
    if device_type:
        query = query.filter_by(device_type=device_type)
    
    if is_physical is not None:
        query = query.filter_by(is_physical=is_physical == 'true')
    
    if platform_id:
        query = query.filter_by(virtualization_platform_id=platform_id)
    
    if tag_id:
        query = query.join(HostTagRelation).filter(HostTagRelation.tag_id == tag_id)
    
    if collection_status:
        query = query.filter_by(collection_status=collection_status)
    
    if source:
        query = query.filter_by(source=source)
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get error messages for failed hosts
    host_dicts = []
    for host in pagination.items:
        host_dict = host.to_dict()
        # Add error_message from latest failed HostDetail if collection_status is 'failed'
        if host.collection_status == 'failed':
            latest_failed_detail = HostDetail.query.filter_by(
                host_id=host.id,
                status='failed'
            ).order_by(HostDetail.collected_at.desc()).first()
            if latest_failed_detail and latest_failed_detail.error_message:
                host_dict['error_message'] = latest_failed_detail.error_message
        host_dicts.append(host_dict)
    
    return jsonify({
        'code': 200,
        'data': host_dicts,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': pagination.total,
            'pages': pagination.pages
        }
    })


@bp.route('/<int:host_id>', methods=['GET'])
@jwt_required()
def get_host(host_id):
    """Get host details"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    
    host_dict = host.to_dict(include_details=True)
    # Add error_message from latest failed HostDetail if collection_status is 'failed'
    if host.collection_status == 'failed':
        latest_failed_detail = HostDetail.query.filter_by(
            host_id=host.id,
            status='failed'
        ).order_by(HostDetail.collected_at.desc()).first()
        if latest_failed_detail and latest_failed_detail.error_message:
            host_dict['error_message'] = latest_failed_detail.error_message
    
    return jsonify({
        'code': 200,
        'data': host_dict
    })


@bp.route('', methods=['POST'])
@jwt_required()
@validate_json(['ip'])
def create_host():
    """Create a new host"""
    data = request.json
    user_id = get_current_user_id()
    
    # Check if host already exists
    existing = Host.query.filter_by(ip=data['ip'], deleted_at=None).first()
    if existing:
        return jsonify({'code': 400, 'message': 'Host with this IP already exists'}), 400
    
    host = Host(
        ip=data['ip'],
        hostname=data.get('hostname'),
        mac=data.get('mac'),
        vendor=data.get('vendor'),
        os_type=data.get('os_type'),
        os_version=data.get('os_version'),
        device_type=data.get('device_type', 'host'),
        is_physical=data.get('is_physical', True),
        virtualization_platform_id=data.get('virtualization_platform_id'),
        created_by=user_id,
    )
    
    db.session.add(host)
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': 'Host created',
        'data': host.to_dict()
    }), 201


@bp.route('/<int:host_id>', methods=['PUT'])
@jwt_required()
def update_host(host_id):
    """Update host"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    data = request.json
    
    # Update fields
    if 'hostname' in data:
        host.hostname = data['hostname']
    if 'mac' in data:
        host.mac = data['mac']
    if 'vendor' in data:
        host.vendor = data['vendor']
    if 'os_type' in data:
        host.os_type = data['os_type']
    if 'os_version' in data:
        host.os_version = data['os_version']
    if 'device_type' in data:
        host.device_type = data['device_type']
    if 'is_physical' in data:
        host.is_physical = data['is_physical']
    if 'virtualization_platform_id' in data:
        host.virtualization_platform_id = data['virtualization_platform_id']
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': 'Host updated',
        'data': host.to_dict()
    })


@bp.route('/<int:host_id>', methods=['DELETE'])
@jwt_required()
def delete_host(host_id):
    """Soft delete host"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    
    host.deleted_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': 'Host deleted'
    })


@bp.route('/batch/delete', methods=['POST'])
@jwt_required()
@validate_json(['host_ids'])
def batch_delete_hosts():
    """Batch delete hosts"""
    data = request.json
    host_ids = data.get('host_ids', [])
    
    if not host_ids:
        return jsonify({
            'code': 400,
            'message': 'No host IDs provided'
        }), 400
    
    deleted_count = 0
    failed_ids = []
    
    for host_id in host_ids:
        try:
            host = Host.query.filter_by(id=host_id, deleted_at=None).first()
            if host:
                host.deleted_at = datetime.utcnow()
                deleted_count += 1
            else:
                failed_ids.append(host_id)
        except Exception as e:
            failed_ids.append(host_id)
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': f'Deleted {deleted_count} hosts',
        'data': {
            'deleted_count': deleted_count,
            'failed_ids': failed_ids
        }
    })


@bp.route('/batch', methods=['POST'])
@jwt_required()
def batch_create_hosts():
    """Batch create/update hosts"""
    data = request.json
    hosts_data = data.get('hosts', [])
    user_id = get_current_user_id()
    
    created = []
    updated = []
    
    for host_data in hosts_data:
        ip = host_data.get('ip')
        if not ip:
            continue
        
        existing = Host.query.filter_by(ip=ip, deleted_at=None).first()
        
        if existing:
            # Update existing
            for key, value in host_data.items():
                if hasattr(existing, key):
                    setattr(existing, key, value)
            updated.append(existing)
        else:
            # Create new
            host = Host(
                ip=ip,
                hostname=host_data.get('hostname'),
                mac=host_data.get('mac'),
                vendor=host_data.get('vendor'),
                os_type=host_data.get('os_type'),
                os_version=host_data.get('os_version'),
                device_type=host_data.get('device_type', 'host'),
                is_physical=host_data.get('is_physical', True),
                created_by=user_id,
            )
            db.session.add(host)
            created.append(host)
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': f'Created {len(created)}, updated {len(updated)} hosts',
        'data': {
            'created': [h.to_dict() for h in created],
            'updated': [h.to_dict() for h in updated]
        }
    })


@bp.route('/<int:host_id>/credentials', methods=['POST'])
@jwt_required()
@validate_json(['username'])
def set_credentials(host_id):
    """Set host credentials"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    data = request.json
    
    credential = HostCredential.query.filter_by(host_id=host_id).first()
    
    if not credential:
        credential = HostCredential(host_id=host_id)
        db.session.add(credential)
    
    credential.username = data['username']
    if 'password' in data:
        credential.password_encrypted = encrypt_password(data['password'])
    if 'ssh_port' in data:
        credential.ssh_port = data['ssh_port']
    if 'key_path' in data:
        credential.key_path = data['key_path']
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': 'Credentials updated',
        'data': credential.to_dict(decrypt_password=True)
    })


@bp.route('/<int:host_id>/credentials', methods=['GET'])
@jwt_required()
def get_credentials(host_id):
    """Get host credentials (with decrypted password)"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    credential = HostCredential.query.filter_by(host_id=host_id).first()
    
    if not credential:
        return jsonify({
            'code': 404,
            'message': 'No credentials found'
        }), 404
    
    return jsonify({
        'code': 200,
        'data': credential.to_dict(decrypt_password=True)
    })


@bp.route('/batch/collect', methods=['POST'])
@jwt_required()
@validate_json(['host_ids'])
def batch_collect():
    """Trigger batch collection"""
    data = request.json
    host_ids = data['host_ids']
    concurrent_limit = data.get('concurrent_limit')
    user_id = get_current_user_id()
    
    # Get hosts and check their source and credentials
    hosts = Host.query.filter(
        Host.id.in_(host_ids),
        Host.deleted_at == None
    ).all()
    
    if not hosts:
        return jsonify({
            'code': 404,
            'message': 'No valid hosts found'
        }), 404
    
    # Separate hosts by source
    platform_hosts = []
    scan_hosts = []
    missing_credentials = []
    
    for host in hosts:
        if host.source == 'platform':
            platform_hosts.append(host)
        elif host.source == 'scan':
            # Check if credentials exist
            credential = HostCredential.query.filter_by(host_id=host.id).first()
            if not credential or not credential.username:
                missing_credentials.append({
                    'id': host.id,
                    'ip': host.ip,
                    'hostname': host.hostname
                })
            else:
                scan_hosts.append(host)
        else:
            # For manual hosts, also check credentials
            credential = HostCredential.query.filter_by(host_id=host.id).first()
            if not credential or not credential.username:
                missing_credentials.append({
                    'id': host.id,
                    'ip': host.ip,
                    'hostname': host.hostname
                })
            else:
                scan_hosts.append(host)
    
    # If there are missing credentials, return error with list
    if missing_credentials:
        return jsonify({
            'code': 400,
            'message': 'Some hosts are missing authentication information',
            'data': {
                'missing_credentials': missing_credentials
            }
        }), 400
    
    # Handle platform hosts separately
    if platform_hosts:
        # Group by platform
        from collections import defaultdict
        platform_groups = defaultdict(list)
        for host in platform_hosts:
            platform_id = host.source_platform_id or host.virtualization_platform_id
            if platform_id:
                platform_groups[platform_id].append(host.id)
        
        # Create platform collection tasks
        platform_tasks = []
        for platform_id, p_host_ids in platform_groups.items():
            from tasks.collector import collect_platform_hosts_task
            # Create a collection task for platform hosts
            task = CollectionTask(
                concurrent_limit=1,  # Platform collection is usually sequential
                status='pending',
                created_by=user_id,
            )
            task.set_host_ids(p_host_ids)
            db.session.add(task)
            db.session.flush()
            platform_tasks.append(task)
            
            # Start async platform collection task
            collect_platform_hosts_task.delay(task.id, platform_id)
    
    # Handle scan/manual hosts with normal collection
    if scan_hosts:
        scan_host_ids = [h.id for h in scan_hosts]
        task = CollectionTask(
            concurrent_limit=concurrent_limit or 5,
            status='pending',
            created_by=user_id,
        )
        task.set_host_ids(scan_host_ids)
        db.session.add(task)
        db.session.flush()
        
        # Start async task
        collect_hosts_task.delay(task.id, concurrent_limit)
    
    db.session.commit()
    
    # Return response
    if platform_hosts and scan_hosts:
        message = f'Created {len(platform_tasks)} platform collection task(s) and 1 normal collection task'
    elif platform_hosts:
        message = f'Created {len(platform_tasks)} platform collection task(s)'
    else:
        message = 'Collection task created'
    
    return jsonify({
        'code': 200,
        'message': message,
        'data': {
            'platform_tasks': [t.to_dict() for t in platform_tasks] if platform_hosts else [],
            'normal_task': task.to_dict() if scan_hosts else None
        }
    }), 201


@bp.route('/<int:host_id>/details', methods=['GET'])
@jwt_required()
def get_host_details(host_id):
    """Get host collection details"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    
    details = HostDetail.query.filter_by(host_id=host_id).order_by(HostDetail.collected_at.desc()).all()
    
    return jsonify({
        'code': 200,
        'data': [detail.to_dict() for detail in details]
    })


@bp.route('/<int:host_id>/relationships', methods=['GET'])
@jwt_required()
def get_relationships(host_id):
    """Get host relationships"""
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    
    outgoing = HostRelationship.query.filter_by(from_host_id=host_id).all()
    incoming = HostRelationship.query.filter_by(to_host_id=host_id).all()
    
    return jsonify({
        'code': 200,
        'data': {
            'outgoing': [r.to_dict() for r in outgoing],
            'incoming': [r.to_dict() for r in incoming]
        }
    })


@bp.route('/batch/credentials', methods=['POST'])
@jwt_required()
@validate_json(['host_ids'])
def batch_update_credentials():
    """Batch update host credentials"""
    data = request.json
    host_ids = data['host_ids']
    credentials = data.get('credentials', {})
    
    if not credentials:
        return jsonify({
            'code': 400,
            'message': 'No credentials provided'
        }), 400
    
    updated = []
    errors = []
    
    for host_id in host_ids:
        try:
            host = Host.query.filter_by(id=host_id, deleted_at=None).first()
            if not host:
                errors.append({'host_id': host_id, 'error': 'Host not found'})
                continue
            
            credential = HostCredential.query.filter_by(host_id=host_id).first()
            if not credential:
                credential = HostCredential(host_id=host_id)
                db.session.add(credential)
            
            if 'username' in credentials:
                credential.username = credentials['username']
            if 'password' in credentials:
                credential.password_encrypted = encrypt_password(credentials['password'])
            if 'ssh_port' in credentials:
                credential.ssh_port = credentials.get('ssh_port', 22)
            if 'key_path' in credentials:
                credential.key_path = credentials.get('key_path')
            
            updated.append(host_id)
            
        except Exception as e:
            errors.append({'host_id': host_id, 'error': str(e)})
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': f'Updated credentials for {len(updated)} hosts, {len(errors)} errors',
        'data': {
            'updated': updated,
            'errors': errors
        }
    })


@bp.route('/batch/update', methods=['POST'])
@jwt_required()
@validate_json(['hosts'])
def batch_update_hosts():
    """Batch update hosts (for online table editing)"""
    data = request.json
    hosts_data = data['hosts']  # List of {id, ...fields to update}
    user_id = get_current_user_id()
    
    updated = []
    errors = []
    
    for host_data in hosts_data:
        try:
            host_id = host_data.get('id')
            if not host_id:
                errors.append({'host': host_data, 'error': 'Missing host id'})
                continue
            
            host = Host.query.filter_by(id=host_id, deleted_at=None).first()
            if not host:
                errors.append({'host': host_data, 'error': f'Host {host_id} not found'})
                continue
            
            # Update allowed fields
            allowed_fields = [
                'hostname', 'mac', 'vendor', 'os_type', 'os_version', 'os_kernel', 'os_bit',
                'boot_type', 'cpu_info', 'cpu_cores', 'memory_total', 'memory_free', 'memory_info',
                'device_type', 'is_physical', 'vt_platform', 'vt_platform_ver'
            ]
            
            for field in allowed_fields:
                if field in host_data:
                    setattr(host, field, host_data[field] or None)
            
            # Update credentials if provided
            if 'credentials' in host_data:
                cred_data = host_data['credentials']
                credential = HostCredential.query.filter_by(host_id=host_id).first()
                
                if not credential:
                    credential = HostCredential(host_id=host_id)
                    db.session.add(credential)
                
                if 'username' in cred_data:
                    credential.username = cred_data['username']
                if 'password' in cred_data:
                    credential.password_encrypted = encrypt_password(cred_data['password'])
                if 'ssh_port' in cred_data:
                    credential.ssh_port = cred_data.get('ssh_port', 22)
                if 'key_path' in cred_data:
                    credential.key_path = cred_data.get('key_path')
            
            updated.append(host)
            
        except Exception as e:
            errors.append({'host': host_data, 'error': str(e)})
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'message': f'Updated {len(updated)} hosts, {len(errors)} errors',
        'data': {
            'updated': [h.to_dict() for h in updated],
            'errors': errors
        }
    })


@bp.route('/tree', methods=['GET'])
@jwt_required()
def get_hosts_tree():
    """Get hosts in tree structure (platform -> ESXi -> VM)"""
    # Only get platform-sourced hosts
    hosts = Host.query.filter(
        Host.source == 'platform',
        Host.deleted_at == None
    ).all()
    
    # Separate ESXi hosts and VMs
    from collections import defaultdict
    platform_tree = defaultdict(lambda: {
        'platform_id': None,
        'platform_name': None,
        'platform_type': None,
        'esxi_hosts': defaultdict(lambda: {
            'esxi_id': None,
            'esxi_name': None,
            'esxi_ip': None,
            'vms': []
        })
    })
    
    # First pass: identify ESXi hosts and VMs
    esxi_hosts = {}  # {esxi_name: host_obj}
    vms = []  # List of (host_obj, esxi_name)
    
    for host in hosts:
        platform_id = host.source_platform_id or host.virtualization_platform_id
        if not platform_id:
            continue
        
        # Get platform info
        from models import VirtualizationPlatform
        platform = VirtualizationPlatform.query.get(platform_id)
        if not platform:
            continue
        
        # For VMware, identify ESXi hosts by os_type
        is_esxi = (host.os_type == 'VMware ESXi')
        
        if is_esxi:
            esxi_name = host.hostname or host.ip
            esxi_hosts[esxi_name] = host
        else:
            # This is a VM, extract ESXi name from vendor field if stored
            esxi_name = None
            if host.vendor and host.vendor.startswith('ESXi: '):
                esxi_name = host.vendor.replace('ESXi: ', '')
            vms.append((host, esxi_name))
    
    # Group VMs by platform and ESXi
    for host, esxi_name in vms:
        platform_id = host.source_platform_id or host.virtualization_platform_id
        if not platform_id:
            continue
        
        from models import VirtualizationPlatform
        platform = VirtualizationPlatform.query.get(platform_id)
        if not platform:
            continue
        
        platform_node = platform_tree[platform_id]
        platform_node['platform_id'] = platform_id
        platform_node['platform_name'] = platform.name
        platform_node['platform_type'] = platform.type
        
        # Use ESXi name if available, otherwise group under "Unknown ESXi"
        esxi_key = esxi_name if esxi_name else 'Unknown ESXi'
        
        platform_node['esxi_hosts'][esxi_key]['esxi_name'] = esxi_key
        platform_node['esxi_hosts'][esxi_key]['vms'].append(host.to_dict())
    
    # Add ESXi hosts to tree
    for esxi_name, esxi_host in esxi_hosts.items():
        platform_id = esxi_host.source_platform_id or esxi_host.virtualization_platform_id
        if not platform_id:
            continue
        
        from models import VirtualizationPlatform
        platform = VirtualizationPlatform.query.get(platform_id)
        if not platform:
            continue
        
        platform_node = platform_tree[platform_id]
        platform_node['platform_id'] = platform_id
        platform_node['platform_name'] = platform.name
        platform_node['platform_type'] = platform.type
        
        esxi_key = esxi_name
        platform_node['esxi_hosts'][esxi_key]['esxi_id'] = esxi_host.id
        platform_node['esxi_hosts'][esxi_key]['esxi_name'] = esxi_name
        platform_node['esxi_hosts'][esxi_key]['esxi_ip'] = esxi_host.ip
    
    # Convert to list format
    result = []
    for platform_id, platform_data in platform_tree.items():
        esxi_list = []
        for esxi_key, esxi_data in platform_data['esxi_hosts'].items():
            esxi_list.append({
                'esxi_id': esxi_data.get('esxi_id'),
                'esxi_name': esxi_data['esxi_name'],
                'esxi_ip': esxi_data.get('esxi_ip'),
                'vms': esxi_data['vms']
            })
        
        result.append({
            'platform_id': platform_data['platform_id'],
            'platform_name': platform_data['platform_name'],
            'platform_type': platform_data['platform_type'],
            'esxi_hosts': esxi_list
        })
    
    return jsonify({
        'code': 200,
        'data': result
    })


@bp.route('/export/csv', methods=['GET'])
@jwt_required()
def export_hosts_csv():
    """Export host data to CSV or generate import template"""
    template_flag = str(request.args.get('template', '')).lower() in ('1', 'true', 'yes', 'template')
    filter_keys = [
        'search', 'search_field', 'os_type', 'device_type', 'collection_status',
        'source', 'is_physical', 'tag_id', 'platform_id'
    ]
    filters = {key: request.args.get(key) for key in filter_keys}
    
    if template_flag:
        prepopulate = (request.args.get('prepopulate') or 'none').lower()
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([column['key'] for column in HOST_IMPORT_TEMPLATE_COLUMNS])
        
        hosts = []
        if prepopulate in ('filtered', 'not_collected'):
            template_filters = {key: value for key, value in filters.items()}
            if prepopulate == 'not_collected':
                template_filters['collection_status'] = 'not_collected'
            query = Host.query.options(*HOST_TEMPLATE_EXPORT_OPTIONS).filter(Host.deleted_at == None)
            query = _apply_host_filters(query, template_filters)
            hosts = query.order_by(Host.id.asc()).all()
        
        for host in hosts:
            credential = host.credentials[0] if host.credentials else None
            row = []
            for column in HOST_IMPORT_TEMPLATE_COLUMNS:
                key = column['key']
                value = ''
                if key == 'ip':
                    value = host.ip or ''
                elif key == 'hostname':
                    value = host.hostname or ''
                elif key == 'mac':
                    value = host.mac or ''
                elif key == 'vendor':
                    value = host.vendor or ''
                elif key == 'os_type':
                    value = host.os_type or ''
                elif key == 'os_version':
                    value = host.os_version or ''
                elif key == 'device_type':
                    value = host.device_type or ''
                elif key == 'is_physical':
                    value = '是' if host.is_physical else '否'
                elif key == 'vt_platform':
                    value = host.vt_platform or ''
                elif key == 'vt_platform_ver':
                    value = host.vt_platform_ver or ''
                elif key == 'username':
                    value = credential.username if credential else ''
                elif key == 'password':
                    value = ''
                elif key == 'ssh_port':
                    value = credential.ssh_port if credential and credential.ssh_port else ''
                elif key == 'key_path':
                    value = credential.key_path if credential and credential.key_path else ''
                elif key == 'tags':
                    value = ','.join(sorted({tag.name for tag in host.tags})) if host.tags else ''
                else:
                    attr = getattr(host, key, None)
                    value = attr if attr is not None else ''
                row.append(value)
            writer.writerow(row)
        
        output.seek(0)
        suffix_map = {
            'not_collected': 'not_collected',
            'filtered': 'filtered',
        }
        suffix = suffix_map.get(prepopulate, 'template')
        filename = f'host_import_{suffix}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
        return Response(
            output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
    
    # Regular export branch
    include_credentials = str(request.args.get('include_credentials', 'false')).lower() == 'true'
    query = Host.query.options(selectinload(Host.credentials)).filter(Host.deleted_at == None)
    query = _apply_host_filters(query, filters)
    hosts = query.order_by(Host.id.asc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    header = [
        'ID', '主机名', 'IP', 'Mac', '厂商', '操作系统类型', '操作系统版本', '操作系统内核',
        '操作系统位数', '启动方式', 'CPU信息', 'CPU核数', '内存(GB)', '剩余内存(GB)',
        '磁盘数量', '磁盘总容量(GB)', '网卡数量', '设备类型', '是否物理机', '虚拟化平台',
        '虚拟化版本', '来源', '采集状态', '最后采集时间'
    ]
    
    if include_credentials:
        header.extend(['用户名', '密码', 'SSH端口', '密钥路径'])
    
    writer.writerow(header)
    
    for host in hosts:
        row = [
            host.id,
            host.hostname or '',
            host.ip or '',
            host.mac or '',
            host.vendor or '',
            host.os_type or '',
            host.os_version or '',
            host.os_kernel or '',
            host.os_bit or '',
            host.boot_type or '',
            host.cpu_info or '',
            host.cpu_cores or 0,
            host.memory_total or 0,
            host.memory_free or 0,
            host.disk_count or 0,
            host.disk_total_size or 0,
            host.network_count or 0,
            host.device_type or '',
            '是' if host.is_physical else '否',
            host.vt_platform or '',
            host.vt_platform_ver or '',
            host.source or 'manual',
            host.collection_status or 'not_collected',
            host.last_collected_at.isoformat() if host.last_collected_at else '',
        ]
        
        if include_credentials:
            credential = host.credentials[0] if host.credentials else None
            if credential:
                password_plain = ''
                if credential.password_encrypted:
                    try:
                        password_plain = decrypt_password(credential.password_encrypted)
                    except Exception:
                        password_plain = ''
                row.extend([
                    credential.username or '',
                    password_plain,
                    credential.ssh_port or 22,
                    credential.key_path or ''
                ])
            else:
                row.extend(['', '', '', ''])
        
        writer.writerow(row)
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename=hosts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        }
    )


def _apply_host_filters(query, filters):
    if not filters:
        return query
    
    search = filters.get('search')
    search_field = filters.get('search_field', 'all')
    os_type = filters.get('os_type')
    device_type = filters.get('device_type')
    tag_id = filters.get('tag_id')
    collection_status = filters.get('collection_status')
    source = filters.get('source')
    is_physical = filters.get('is_physical')
    platform_id = filters.get('platform_id')
    
    if search:
        if search_field == 'ip':
            query = query.filter(Host.ip.contains(search))
        elif search_field == 'hostname':
            query = query.filter(Host.hostname.contains(search))
        elif search_field == 'mac':
            query = query.filter(Host.mac.contains(search))
        elif search_field == 'vendor':
            query = query.filter(Host.vendor.contains(search))
        else:
            query = query.filter(
                or_(
                    Host.hostname.contains(search),
                    Host.ip.contains(search),
                    Host.mac.contains(search),
                    Host.vendor.contains(search)
                )
            )
    
    if os_type:
        query = query.filter_by(os_type=os_type)
    
    if device_type:
        query = query.filter_by(device_type=device_type)
    
    if collection_status:
        query = query.filter_by(collection_status=collection_status)
    
    if source:
        query = query.filter_by(source=source)
    
    if platform_id:
        try:
            platform_id = int(platform_id)
            query = query.filter(
                or_(
                    Host.source_platform_id == platform_id,
                    Host.virtualization_platform_id == platform_id
                )
            )
        except (TypeError, ValueError):
            pass
    
    if is_physical is not None and is_physical != '':
        if isinstance(is_physical, str):
            is_physical_bool = is_physical.lower() == 'true'
        else:
            is_physical_bool = bool(is_physical)
        query = query.filter_by(is_physical=is_physical_bool)
    
    if tag_id:
        try:
            tag_id_int = int(tag_id)
            query = query.join(HostTagRelation).filter(HostTagRelation.tag_id == tag_id_int)
        except (TypeError, ValueError):
            pass
    
    return query


def _resolve_esxi_name(host):
    if host.vendor and host.vendor.startswith('ESXi: '):
        return host.vendor.replace('ESXi: ', '')
    if host.platform and host.platform.name:
        return host.platform.name
    return ''


def _format_disks(host):
    entries = []
    for disk in host.disks:
        size_gb = _bytes_to_gb(disk.size)
        entries.append(f"{disk.device or ''}|{size_gb}|{disk.vendor or ''}|{disk.model or ''}")
    return '\n'.join(entries)


def _format_partitions(host):
    entries = []
    for part in host.partitions:
        total_gb = _bytes_to_gb(part.size_total)
        available_gb = _bytes_to_gb(part.size_available)
        if part.size_available_ratio is not None:
            usage_ratio = round((1 - part.size_available_ratio) * 100, 2)
        else:
            usage_ratio = ''
        entries.append(f"{part.device or ''}|{total_gb}|{available_gb}|{usage_ratio}|{part.fstype or ''}")
    return '\n'.join(entries)


def _format_networks(host):
    entries = []
    for nic in host.network_interfaces:
        if nic.active is True:
            active_str = 'True'
        elif nic.active is False:
            active_str = 'False'
        else:
            active_str = ''
        entries.append(
            f"{nic.interface or ''}|{nic.macaddress or ''}|{active_str}|"
            f"{nic.mtu or ''}|{nic.speed or ''}|{nic.ipv4_address or ''}|"
            f"{nic.ipv4_netmask or ''}|{nic.ipv4_network or ''}|"
            f"{nic.ipv4_broadcast or ''}|{nic.gateway or ''}"
        )
    return '\n'.join(entries)


def _get_field_value(field_key, host):
    if field_key == 'platform_type':
        return 'Physical' if host.is_physical else (host.vt_platform or 'Virtual')
    if field_key == 'hostname':
        return host.hostname or ''
    if field_key == 'vmware_host':
        return host.platform.name if host.platform and host.platform.name else ''
    if field_key == 'ip':
        return host.ip or ''
    if field_key == 'mac':
        return host.mac or ''
    if field_key == 'os_type':
        return host.os_type or ''
    if field_key == 'os_version':
        return host.os_version or ''
    if field_key == 'os_bit':
        return host.os_bit or ''
    if field_key == 'os_kernel':
        return host.os_kernel or ''
    if field_key == 'boot_type':
        return host.boot_type or ''
    if field_key == 'cpu_info':
        return host.cpu_info or ''
    if field_key == 'cpu_cores':
        return host.cpu_cores or ''
    if field_key == 'memory_label':
        return host.memory_info or ''
    if field_key == 'memory_total_gb':
        return host.memory_total or ''
    if field_key == 'memory_free_gb':
        return host.memory_free or ''
    if field_key == 'disk_count':
        return host.disk_count or ''
    if field_key == 'disk_total_size_gb':
        return host.disk_total_size or ''
    if field_key == 'disks':
        return _format_disks(host)
    if field_key == 'partitions':
        return _format_partitions(host)
    if field_key == 'network_count':
        if host.network_count is not None:
            return host.network_count
        return len(host.network_interfaces or [])
    if field_key == 'networks':
        return _format_networks(host)
    if field_key == 'vt_platform':
        return host.vt_platform or ''
    if field_key == 'vt_platform_ver':
        return host.vt_platform_ver or ''
    if field_key == 'esxi_host':
        return _resolve_esxi_name(host)
    if field_key == 'collection_status':
        return host.collection_status or ''
    if field_key == 'last_collected_at':
        return host.last_collected_at.isoformat() if host.last_collected_at else ''
    return ''


@bp.route('/export/excel', methods=['POST'])
@jwt_required()
def export_hosts_excel():
    """Export hosts as Excel workbook with customizable fields"""
    data = request.json or {}
    host_ids = data.get('host_ids', [])
    select_all = data.get('select_all', False)
    filters = data.get('filters', {}) if select_all else {}
    requested_fields = data.get('fields')
    template_name = data.get('template')
    
    if requested_fields and isinstance(requested_fields, list):
        requested_fields = [field for field in requested_fields if field in FIELD_DEFINITIONS]
    else:
        requested_fields = None
    
    if template_name and template_name in EXPORT_TEMPLATES:
        template_fields = EXPORT_TEMPLATES[template_name]['fields']
    else:
        template_fields = EXPORT_TEMPLATES[DEFAULT_EXPORT_TEMPLATE]['fields']
    
    fields = requested_fields or template_fields
    if not fields:
        fields = EXPORT_TEMPLATES[DEFAULT_EXPORT_TEMPLATE]['fields']
    
    seen_fields = set()
    filtered_fields = []
    for field in fields:
        if field in FIELD_DEFINITIONS and field not in seen_fields:
            filtered_fields.append(field)
            seen_fields.add(field)
    fields = filtered_fields
    
    if not fields:
        return jsonify({
            'code': 400,
            'message': 'No valid fields selected for export'
        }), 400
    
    if select_all:
        query = Host.query.options(*HOST_EXPORT_OPTIONS).filter(Host.deleted_at == None)
        query = _apply_host_filters(query, filters)
        ordered_hosts = query.order_by(Host.id.asc()).all()
    else:
        if not isinstance(host_ids, list) or len(host_ids) == 0:
            return jsonify({
                'code': 400,
                'message': 'host_ids must be a non-empty list'
            }), 400
        
        normalized_ids = []
        for host_id in host_ids:
            try:
                normalized_ids.append(int(host_id))
            except (TypeError, ValueError):
                continue
        
        if not normalized_ids:
            return jsonify({
                'code': 400,
                'message': 'No valid host IDs provided'
            }), 400
        
        hosts_query = Host.query.options(*HOST_EXPORT_OPTIONS).filter(
            Host.id.in_(normalized_ids),
            Host.deleted_at == None
        ).all()
        
        host_map = {host.id: host for host in hosts_query}
        ordered_hosts = [host_map[host_id] for host_id in normalized_ids if host_id in host_map]
    
    if not ordered_hosts:
        return jsonify({
            'code': 404,
            'message': 'No hosts found for export'
        }), 404
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hosts'
    
    headers = [FIELD_DEFINITIONS[field]['header'] for field in fields]
    ws.append(headers)
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 22
    
    for index, field in enumerate(fields, start=1):
        width = FIELD_DEFINITIONS[field].get('width', 18)
        ws.column_dimensions[get_column_letter(index)].width = width
    
    for host in ordered_hosts:
        row = [_get_field_value(field, host) for field in fields]
        ws.append(row)
        
        current_row = ws.max_row
        for col_index, field in enumerate(fields, start=1):
            cell = ws.cell(row=current_row, column=col_index)
            wrap_text = FIELD_DEFINITIONS[field].get('wrap', False)
            align = FIELD_DEFINITIONS[field].get('align', 'left')
            cell.alignment = Alignment(
                wrap_text=wrap_text,
                vertical='top',
                horizontal='center' if align == 'center' else 'left'
            )
            cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"hosts_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/export/templates', methods=['GET'])
@jwt_required()
def get_host_export_templates():
    """Return available export templates and field metadata"""
    fields_payload = {
        key: {
            'label': definition['label'],
            'label_key': definition['label_key'],
            'category': definition['category'],
            'category_label_key': FIELD_CATEGORIES[definition['category']]['label_key'],
        }
        for key, definition in FIELD_DEFINITIONS.items()
    }
    
    categories_payload = [
        {
            'id': category_id,
            'label': category_data['label'],
            'label_key': category_data['label_key'],
        }
        for category_id, category_data in FIELD_CATEGORIES.items()
    ]
    
    templates_payload = [
        {
            'id': template_key,
            'name': template_data['name'],
            'name_key': template_data['name_key'],
            'description': template_data.get('description'),
            'description_key': template_data.get('description_key'),
            'fields': template_data['fields'],
        }
        for template_key, template_data in EXPORT_TEMPLATES.items()
    ]
    
    return jsonify({
        'code': 200,
        'data': {
            'fields': fields_payload,
            'categories': categories_payload,
            'templates': templates_payload,
            'default_template': DEFAULT_EXPORT_TEMPLATE,
            'all_fields': list(FIELD_DEFINITIONS.keys())
        }
    })


@bp.route('/<int:host_id>/export', methods=['GET'])
@jwt_required()
def export_host_data(host_id):
    """Export host collection data"""
    from flask import Response
    import yaml as yaml_lib
    
    host = Host.query.filter_by(id=host_id, deleted_at=None).first_or_404()
    
    # Get latest detail
    detail = HostDetail.query.filter_by(host_id=host_id).order_by(HostDetail.collected_at.desc()).first()
    
    if not detail:
        return jsonify({
            'code': 404,
            'message': 'No collection data found'
        }), 404
    
    # Get export format (json or yaml)
    export_format = request.args.get('format', 'json').lower()
    
    # Prepare export data
    export_data = {
        'host': host.to_dict(include_details=True),
        'collection_detail': detail.to_dict(),
        'exported_at': datetime.utcnow().isoformat()
    }
    
    if export_format == 'yaml':
        # Export as YAML
        yaml_content = yaml_lib.dump(export_data, default_flow_style=False, allow_unicode=True, sort_keys=False)
        return Response(
            yaml_content,
            mimetype='text/yaml',
            headers={
                'Content-Disposition': f'attachment; filename=host_{host.id}_{host.hostname or host.ip}_export.yaml'
            }
        )
    else:
        # Export as JSON (default)
        return Response(
            json.dumps(export_data, ensure_ascii=False, indent=2, default=str),
            mimetype='application/json',
            headers={
                'Content-Disposition': f'attachment; filename=host_{host.id}_{host.hostname or host.ip}_export.json'
            }
        )

